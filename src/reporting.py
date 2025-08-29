from typing import Dict, Sequence
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from pyrit.memory import MemoryInterface

from data_types import (
    PromptResult
)
from logging_handler import logging
from config_handler import load_test_config

from datetime import datetime
import os
from pathlib import Path
import csv
import json

# Use pathlib.Path to handle paths in an OS-independent way
OUTPUTS_DIR = Path(load_test_config()['output']['base_dir'])
OUTPUTS_DIR.mkdir(parents=True, exist_ok=True) # Create output directory if it doesn't exist

PROMPT_EVALS_DIR = OUTPUTS_DIR / load_test_config()['output']['prompt_eval_dir']
PROMPT_EVALS_DIR.mkdir(parents=True, exist_ok=True)

REPORTS_DIR = OUTPUTS_DIR / load_test_config()['output']['report_dir']
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

_FILE_PROMPT_RESULT_PREFIX='prompt_results'

def save_prompt_results_to_csv(
        *,
        results: Sequence[PromptResult],
        flat: bool=True, # slim business representation
        compact: bool=False, # it has effect only if flat=False
        results_subfolder: str='',
        test_name: str=f"test_{datetime.now().strftime('%d%m%Y_%H%M%S')}"
    ) -> None:
    '''
    flat: bool -> produces a slim dump of the results with just what's needed for business.
    compact: bool -> whether results should be complete (extended) or reduced. It has effect only if flat=False. To be use for getting "under the hood" info of each prompt evaluation.
    '''
    if not results:
        print("No data to write.")
        return

    dict_rows = {}
    if flat:
        dict_rows = [result.to_dict_flat() for result in results]
    else:
        result_mapper = PromptResult.to_dict_reduced if compact else PromptResult.to_dict_extended
        dict_rows = [result_mapper(result) for result in results]

        # Convert JSON-like fields into proper JSON strings
        for row in dict_rows:
            if isinstance(row.get("scores"), (dict, list)):
                row["scores"] = json.dumps(row["scores"], ensure_ascii=False)

    base_dir = PROMPT_EVALS_DIR / results_subfolder if results_subfolder else PROMPT_EVALS_DIR
    # Create output directory if it doesn't exist
    base_dir.mkdir(parents=True, exist_ok=True)
    file_path= base_dir / f"{_FILE_PROMPT_RESULT_PREFIX}_{test_name}.csv"

    # Write to CSV
    with file_path.open(mode='w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=dict_rows[0].keys())
        writer.writeheader()
        writer.writerows(dict_rows)

def load_results_from_csv(*,
    file_name: str,
    compact: bool=False,
    output_subfolder: str,
    ) -> Sequence[PromptResult]:
    f_suffix = Path(file_name).suffix
    if f_suffix and f_suffix != '.csv':
        raise ValueError(f"{f_suffix} extension is not valid. Only results from csv files can be loaded")
    fname = file_name if f_suffix else f"{file_name}.csv"
    out_folder_path = PROMPT_EVALS_DIR if not output_subfolder else PROMPT_EVALS_DIR / output_subfolder 
    file_path= out_folder_path / fname
    
    results: Sequence[PromptResult] = []
    with file_path.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            results.append(PromptResult.from_dict(row, compact=compact))
    return results

def dump_to_json(*, memory: MemoryInterface) -> None:
    memory.export_conversations(file_path=Path(OUTPUTS_DIR)/"conversation.json")

# ---  ---
def apply_conditional_formatting(path: str):
    """
    Utility for optional conditional formatting
    Apply conditional formatting to perc_jailbreaks columns in an Excel file.
    """
    from openpyxl import load_workbook
    wb = load_workbook(path)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Find column of perc_jailbreaks (header row is 1)
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "perc_jailbreaks":
                col_letter = ws.cell(row=1, column=col).column_letter
                cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"

                # >1% orange
                ws.conditional_formatting.add(
                    cell_range,
                    CellIsRule(operator="greaterThan", formula=["0.01"],
                               fill=PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"))
                )
                # >2% red
                ws.conditional_formatting.add(
                    cell_range,
                    CellIsRule(operator="greaterThan", formula=["0.02"],
                               fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"))
                )
    wb.save(path)
